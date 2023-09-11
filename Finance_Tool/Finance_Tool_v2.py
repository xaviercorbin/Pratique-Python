import os
import sqlite3
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox, ttk

import customtkinter
import openpyxl
import pandas as pd
import yfinance as yf

# Beaucoup de import, mais je compte les utiliser pour des futurs trucs. Par contre, je pense qu'ils en a des inutiles qui se répètent

customtkinter.set_appearance_mode('system')
customtkinter.set_default_color_theme('dark-blue')
mainroot = customtkinter.CTk()
screen_width = mainroot.winfo_screenwidth()
screen_height = mainroot.winfo_screenheight()

# Pour essayer de centrer la fenêtre au centre
x = int(screen_width/2 - 500/2)
y = int(screen_height/2 - 350/2)
mainroot.geometry(f"500x400+{x}+{y}")
mainroot.title("Selection of Tools")


def historic_prices():
    print("------User have selected : Historic Price------")
    frame_historic_prices = customtkinter.CTk()
    frame_historic_prices.title('Historic Price')
    frame_historic_prices.geometry("550x150")
    option_var = customtkinter.StringVar(value='1mo')

    def period_option_menu(choice):
        period_textbox.delete(1.0, END)
        if choice == '1mo':
            period_textbox.insert(END, '1mo')
        if choice == '3mo':
            period_textbox.insert(END, '3mo')
        if choice == '6mo':
            period_textbox.insert(END, '6mo')
        if choice == 'ytd':
            period_textbox.insert(END, 'ytd')
        if choice == '1y':
            period_textbox.insert(END, '1y')
        if choice == '2y':
            period_textbox.insert(END, '2y')
        if choice == '5y':
            period_textbox.insert(END, '5y')
        if choice == '10y':
            period_textbox.insert(END, '10y')
        elif choice == 'max':
            period_textbox.insert(END, 'max')

    menu_period = customtkinter.CTkOptionMenu(master=frame_historic_prices, values=[
                                              '1mo', '3mo', '6mo', 'ytd', '1y', '2y', '5y', '10y', 'max'], command=period_option_menu, variable=option_var)
    menu_period.grid(row=4, column=3, sticky='W')

    def results_of_historic_prices():
        print('Research of historic prices for the ticker : ' +
              entry_ticker_entry.get())
        frame_results_of_historic_prices = tk.Toplevel()
        frame_results_of_historic_prices.title(
            'Results of the historic price of : ' + entry_ticker_entry.get())

        # Creation of the treeview
        historic_prices_tree = ttk.Treeview(frame_results_of_historic_prices, columns=(
            'Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Dividends', 'Stock Splits'))
        style = ttk.Style()
        style.theme_use("clam")

        # Set Column Headings
        historic_prices_tree.heading(
            '#0', text='Ticker', anchor='center')
        historic_prices_tree.heading(
            '#1', text='Date', anchor='center')
        historic_prices_tree.heading(
            '#2', text='Open', anchor='center')
        historic_prices_tree.heading(
            '#3', text='High', anchor='center')
        historic_prices_tree.heading('#4', text='Low', anchor='center')
        historic_prices_tree.heading(
            '#5', text='Close', anchor='center')
        historic_prices_tree.heading(
            '#6', text='Volume', anchor='center')
        historic_prices_tree.heading(
            '#7', text='Dividends', anchor='center')
        historic_prices_tree.heading(
            '#8', text='Stock Splits', anchor='center')

        # Set Column Widths
        historic_prices_tree.column('#1', width=100)
        historic_prices_tree.column('#2', width=100)
        historic_prices_tree.column('#3', width=100)
        historic_prices_tree.column('#4', width=100)
        historic_prices_tree.column('#5', width=100)
        historic_prices_tree.column('#6', width=100)
        historic_prices_tree.column('#7', width=100)
        historic_prices_tree.column('#8', width=100)

        # Insert the Data in the Treeview
        ticker_symbol = entry_ticker_entry.get()
        ticker = yf.Ticker(ticker_symbol)
        # select the period
        history = ticker.history(period=period_textbox.get('1.0', 'end-1c'))
        for i, row in history.iterrows():
            open_price = round(row['Open'], 2)
            high_price = round(row['High'], 2)
            low_price = round(row['Low'], 2)
            close_price = round(row['Close'], 2)
            volume = round(row['Volume'], 2)
            dividends = round(row['Dividends'], 2)
            stock_splits = round(row['Stock Splits'], 2)

            # insert the data into the treeview
            historic_prices_tree.insert('', 'end', text=ticker.ticker, values=(i.date(
            ), f"{open_price:^10}", f"{high_price:^10}", f"{low_price:^10}", f"{close_price:^10}", f"{volume:^10}", f"{dividends:^10}", f"{stock_splits:^10}"))

        # pack the treeview into the root window
        historic_prices_tree.pack(
            fill='both', expand=True, padx=10, side='left')

        # create a vertical scrollbar for the treeview
        vsb = ttk.Scrollbar(frame_results_of_historic_prices, orient="vertical",
                            command=historic_prices_tree.yview)
        vsb.pack(side='left', fill='y')

        # configure the treeview to use the scrollbar
        historic_prices_tree.configure(yscrollcommand=vsb.set)

        # set the height of the treeviw to show more rows
        historic_prices_tree['height'] = 20

        def export_to_excel(historic_prices_tree):
            wb = openpyxl.Workbook()
            ws = wb.active

            # write column headings
            for col, heading in enumerate(historic_prices_tree["columns"]):
                ws.cell(row=1, column=col+1,
                        value=historic_prices_tree.heading(heading)["text"])

            # write data
            for row, item in enumerate(historic_prices_tree.get_children()):
                for col, value in enumerate(historic_prices_tree.item(item)["values"]):
                    ws.cell(row=row+2, column=col+1, value=value)

            # save the workbook
            wb.save(ticker_symbol+"_treeview_data.xlsx")
            print("Exported to Excel")

        export_button = tk.Button(frame_results_of_historic_prices, text="Export to Excel",
                                  command=lambda: export_to_excel(historic_prices_tree))
        export_button.pack(side='right', padx=10, pady=10)

        frame_results_of_historic_prices.mainloop()

    # Ticker entry section
    label_title_ticker = customtkinter.CTkLabel(
        master=frame_historic_prices, text='Historic Prices', font=("Arial", 24))
    label_title_ticker.grid(row=1, column=1, columnspan=2)

    label_ticker_entry = customtkinter.CTkLabel(
        master=frame_historic_prices, text='Enter the stock ticker : ', font=("Arial", 18))
    label_ticker_entry.grid(row=3, column=3, sticky='W')

    entry_ticker_entry = customtkinter.CTkEntry(
        master=frame_historic_prices, width=200)
    entry_ticker_entry.grid(row=3, column=4, sticky='W')
    entry_ticker_entry.focus()

    period_textbox = customtkinter.CTkTextbox(
        master=frame_historic_prices, width=50, height=20)
    period_textbox.grid(row=4, column=4, sticky='W')

    button_search_historic_prices = customtkinter.CTkButton(
        master=frame_historic_prices, text='Search', command=results_of_historic_prices)
    button_search_historic_prices.grid(row=5, column=3, sticky='W')

    frame_historic_prices.mainloop()


def calculation_of_shares_to_buy():
    print('------alculation of shares to buy selected------')

    '''
    - find the stock you want to buy
    - retreive its price per shares
    - input budget
    - input commissions
    - ...
    '''


def new_statement_policy_made():
    print('------New statement policy made------')
    frame_new_statement_policy_made = customtkinter.CTk()
    frame_new_statement_policy_made.geometry("800x70")
    frame_new_statement_policy_made.title('Statement Policy')
    customtkinter.set_default_color_theme('green')
    
    frame_new_statement_policy_made.mainloop()
    
    
    


def building_a_portfolio():
    print('Building a portfolio was selected')
    frame_building_a_portfolio_0 = customtkinter.CTk()
    frame_building_a_portfolio_0.geometry('200x100')
    frame_building_a_portfolio_0.title('Building your portfolio')
    customtkinter.set_default_color_theme('dark-blue')

    def building_new_portfolio():
        print('------Building a new portfolio was selected------')
        # Forms pour avoir les détails du user nécessaire à la construction du portefeuille
        frame_building_new_portfolio = customtkinter.CTk()
        frame_building_new_portfolio.geometry('550x300')
        frame_building_new_portfolio.title('Investment Policy Statement')

        # Font for the form
        font1 = ('Arial', 16)

        def submit_form_button():
            print('------summit form button was pressed------')
            Last_Name = last_name_input.get()
            Name = name_input.get()
            Age = age_input.get()
            Years_Before_Retirement = retirement_input.get()
            Monthly_Icome = income_input.get()
            Monthly_Expenses = expenses_input.get()
            Total_Investment = investments_input.get()
            print(Last_Name, Name, Age, Years_Before_Retirement, Monthly_Icome, Monthly_Expenses, Total_Investment)

            paramètre = [Last_Name, Name, Age, Years_Before_Retirement, Monthly_Icome, Monthly_Expenses, Total_Investment]

            if all(paramètre):
                conn = sqlite3.connect('users.db')
                c = conn.cursor()

                # Creating the table with new fields
                c.execute('''CREATE TABLE IF NOT EXISTS Retirement_Plan (
                    Last_Name TEXT, 
                    Name TEXT, 
                    Age TEXT, 
                    Years_Before_Retirement TEXT, 
                    Monthly_Income TEXT, 
                    Monthly_Expenses TEXT, 
                    Total_Investments_Value TEXT
                )''')

                # Inserting values into the table
                c.execute("INSERT INTO Retirement_Plan VALUES (?,?,?,?,?,?,?)",
                        (last_name_input.get(), name_input.get(), age_input.get(), retirement_input.get(), income_input.get(), expenses_input.get(), investments_input.get()))

                conn.commit()
                conn.close()

                frame_building_new_portfolio.destroy()
                frame_building_a_portfolio_0.destroy()

                new_statement_policy_made()

            else:
                print('Please fill the form')
                #######################################################################################################
                # Faire la messagebox qui retourne après au investment policy
                #######################################################################################################


        # Last Name
        last_name_label = customtkinter.CTkLabel(
            master=frame_building_new_portfolio, text='Last Name', font=font1).grid(column=0, row=0)
        last_name_input = customtkinter.CTkEntry(
            master=frame_building_new_portfolio, width=200)
        last_name_input.grid(column=1, row=0)

        # Name
        name_label = customtkinter.CTkLabel(
            master=frame_building_new_portfolio, text='Name', font=font1).grid(column=0, row=1)
        name_input = customtkinter.CTkEntry(
            master=frame_building_new_portfolio, width=200)
        name_input.grid(column=1, row=1)

        # Age
        age_label = customtkinter.CTkLabel(
            master=frame_building_new_portfolio, text='Age', font=font1).grid(column=0, row=2)
        age_input = customtkinter.CTkEntry(
            master=frame_building_new_portfolio, width=200)
        age_input.grid(column=1, row=2)

        # Years before retirement
        retirement_label = customtkinter.CTkLabel(
            master=frame_building_new_portfolio, text='How many years before retirement', font=font1).grid(column=0, row=3)
        retirement_input = customtkinter.CTkEntry(
            master=frame_building_new_portfolio, width=200)
        retirement_input.grid(column=1, row=3)

        # Monthly Income
        income_label = customtkinter.CTkLabel(
            master=frame_building_new_portfolio, text='Monthly Income', font=font1).grid(column=0, row=4)
        income_input = customtkinter.CTkEntry(
            master=frame_building_new_portfolio, width=200)
        income_input.grid(column=1, row=4)

        # Monthly Expenses
        expenses_label = customtkinter.CTkLabel(
            master=frame_building_new_portfolio, text='Monthly Expenses', font=font1).grid(column=0, row=5)
        expenses_input = customtkinter.CTkEntry(
            master=frame_building_new_portfolio, width=200)
        expenses_input.grid(column=1, row=5)

        # Value of total investments
        investments_label = customtkinter.CTkLabel(
            master=frame_building_new_portfolio, text='Value of total investments', font=font1).grid(column=0, row=6)
        investments_input = customtkinter.CTkEntry(
            master=frame_building_new_portfolio, width=200)
        investments_input.grid(column=1, row=6)

        # Submit Button
        customtkinter.CTkButton(master=frame_building_new_portfolio,
                                text='Submit', command=submit_form_button).grid(column=0, row=8)



        frame_building_new_portfolio.mainloop()

    def view_existent_portfolio():
        print('------View existent portfolio was selected------')
        
        frame_view_existent_portfolio = customtkinter.CTk()
        frame_view_existent_portfolio.geometry('300x200')
        frame_view_existent_portfolio.title('View Existent Portfolio')

        def view_statement_policy():
            print("Statement Policy selected")

            # Create a window to display list of names
            frame_view_names = customtkinter.CTk()
            frame_view_names.geometry('300x400')
            frame_view_names.title('Select a Person')

            # Retrieve names from the database
            conn = sqlite3.connect('users.db')
            c = conn.cursor()
            c.execute("SELECT Last_Name, Name FROM Retirement_Plan")
            names = c.fetchall()
            conn.close()

            # Create a listbox to display names
            names_listbox = tk.Listbox(master=frame_view_names)
            for last, first in names:
                names_listbox.insert('end', f"{last}, {first}")
            names_listbox.pack(pady=10, padx=10, fill='both', expand=True)

            def on_name_selected(event):
                index = names_listbox.curselection()[0]
                last_name, name = names[index]

                # Retrieve full details of selected person from the database
                conn = sqlite3.connect('users.db')
                c = conn.cursor()
                c.execute("SELECT * FROM Retirement_Plan WHERE Last_Name=? AND Name=?", (last_name, name))
                person_details = c.fetchone()
                conn.close()

                # Create a window to display details
                frame_details = customtkinter.CTk()
                frame_details.geometry('500x300')
                frame_details.title(f"{last_name}, {name}'s Details")

                customtkinter.CTkLabel(master=frame_details, text=f"{last_name}, {name}", font=('Arial', 16, 'bold')).grid(row=0, column=0, columnspan=2, sticky='w', padx=10, pady=5)

                details_keys = ['Age', 'Years Before Retirement', 'Monthly Income', 'Monthly Expenses', 'Total Investments Value']

                details_labels = {}
                details_entries = {}

                for i, key in enumerate(details_keys, start=1):
                    label = customtkinter.CTkLabel(master=frame_details, text=f"{key}: {person_details[2 + i - 1]}")
                    label.grid(row=i, column=0, padx=10, pady=5)
                    details_labels[key] = label

                    entry = customtkinter.CTkEntry(master=frame_details)
                    entry.insert(0, person_details[2 + i - 1])
                    details_entries[key] = entry

                def switch_to_edit_mode():
                    for key, label in details_labels.items():
                        label.grid_forget()
                        details_entries[key].grid(row=details_keys.index(key) + 1, column=1, padx=10, pady=5)

                    save_button.grid(row=6, column=0, columnspan=2)
                    edit_button.grid_forget()

                def save_edits():
                    new_details = {
                        key: details_entries[key].get() for key in details_keys
                    }

                    # Update database
                    conn = sqlite3.connect('users.db')
                    c = conn.cursor()
                    c.execute("""
                        UPDATE Retirement_Plan 
                        SET Age = ?, Years_Before_Retirement = ?, Monthly_Income = ?, Monthly_Expenses = ?, Total_Investments_Value = ?
                        WHERE Last_Name = ? AND Name = ?""",
                        (new_details['Age'], new_details['Years Before Retirement'], new_details['Monthly Income'],
                        new_details['Monthly Expenses'], new_details['Total Investments Value'], last_name, name))
                    conn.commit()
                    conn.close()

                    # Switch back to label mode
                    for key, entry in details_entries.items():
                        entry.grid_forget()
                        details_labels[key].configure(text=f"{key}: {new_details[key]}")
                        details_labels[key].grid(row=details_keys.index(key) + 1, column=0, padx=10, pady=5)

                    edit_button.grid(row=6, column=0, columnspan=2)
                    save_button.grid_forget()

                edit_button = customtkinter.CTkButton(master=frame_details, text="Edit", command=switch_to_edit_mode)
                edit_button.grid(row=6, column=0, columnspan=2)

                save_button = customtkinter.CTkButton(master=frame_details, text="Save", command=save_edits)

                frame_details.mainloop()

            # Bind selection event
            names_listbox.bind('<<ListboxSelect>>', on_name_selected)

            frame_view_names.mainloop()



        def view_portfolio_person():
            print("Portfolio of a person selected")
            # Implement the action to view the portfolio of a person
            # For example, you can retrieve the data from the database and display it.
            pass

        button_statement_policy = customtkinter.CTkButton(
            master=frame_view_existent_portfolio, text='View Statement Policy', command=view_statement_policy)
        button_statement_policy.pack(pady=10, padx=10)

        button_portfolio_person = customtkinter.CTkButton(
            master=frame_view_existent_portfolio, text='View Portfolio of a Person', command=view_portfolio_person)
        button_portfolio_person.pack(pady=10, padx=10)

        frame_view_existent_portfolio.mainloop()
        

    button__new_portfolio = customtkinter.CTkButton(
        master=frame_building_a_portfolio_0, text='New Portfolio', command=building_new_portfolio)
    button__new_portfolio.pack(pady=10, padx=10)

    button_existent_portfolio = customtkinter.CTkButton(
        master=frame_building_a_portfolio_0, text='Existent portfolio', command=view_existent_portfolio)
    button_existent_portfolio.pack(pady=10, padx=10)


mainframe = customtkinter.CTkFrame(master=mainroot)
mainframe._corner_radius = 40
mainframe.pack(pady=20, padx=60, fill='both', expand=True)

title_mainframe = customtkinter.CTkLabel(
    master=mainframe, text='Financial Tools', font=('Arial', 28))
title_mainframe.pack(pady=12, padx=10)

button_historic_price = customtkinter.CTkButton(
    master=mainframe, text='Historic prices', command=historic_prices)
button_historic_price.pack(pady=12, padx=10)

button_building_portfolio = customtkinter.CTkButton(
    master=mainframe, text='Building your portfolio (soon)', command=building_a_portfolio)
button_building_portfolio.pack(pady=12, padx=10)

button_calculation_of_shares_to_buy = customtkinter.CTkButton(
    master=mainframe, text='Calcul the number of share to buy (soon)', command=calculation_of_shares_to_buy)
button_calculation_of_shares_to_buy.pack(pady=12, padx=10)


mainroot.mainloop()