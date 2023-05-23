import os
import random
import sqlite3
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter import filedialog, messagebox, ttk

import customtkinter
import openpyxl
import pandas as pd
import yfinance as yf

customtkinter.set_appearance_mode("system")
customtkinter.set_default_color_theme("dark-blue")
root = customtkinter.CTk()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

x = int(screen_width/2 - 500/2)
y = int(screen_height/2 - 350/2)
root.geometry(f"500x400+{x}+{y}")
root.title("Login Page")

# J'utilisais ça pour me connecter plus rapidement, question de bypass le login rapidement
def load_credentials():
    if os.path.exists("credentials.txt"):
        with open("credentials.txt", "r") as f:
            lines = f.readlines()
            if len(lines) == 2:
                username = lines[0].strip()
                password = lines[1].strip()
                entry1.insert(0, username)
                entry2.insert(0, password)
                checkbox_var.set(True)


def login():
    print("Login")
    username = entry1.get()
    password = entry2.get()
    remember_me = checkbox_var.get()

    conn = sqlite3.connect('users.db')
    c = conn.cursor()

    c.execute(
        '''SELECT * FROM users WHERE username=? AND password=?''', (username, password))
    data = c.fetchone()

    # Checking if user exists in the database or not
    if data is None:
        messagebox.showwarning(
            "Invalid Credentials!", "Username or Password Incorrect!")
        print("Invalid Credentials")
    else:
        print("Successfully login")
        messagebox.showinfo("Success", f"Welcome {data[1]} {data[2]}!")
        root.destroy()
        open_new_interface()


def forgot_credentials():
    print("Trying to retreive username or password")
    forgot_password_option_frame = customtkinter.CTk()
    forgot_password_option_frame.title("Forgot username or password")
    screen_width = forgot_password_option_frame.winfo_screenmmwidth()
    screen_height = forgot_password_option_frame.winfo_screenheight()
    x = int(screen_width/2 - 500/2)
    y = int(screen_height/2 - 350/2)
    forgot_password_option_frame.geometry(f"400x200+{x}+{y}")
    forgot_password_option_frame.grid_columnconfigure(1, weight=1)

    def forgot_username():
        print("forgot username")
        forgot_password_option_frame.destroy()
        forgot_username_frame = customtkinter.CTk()
        forgot_username_frame.title("Forgot Username")
        forgot_username_frame.geometry("400x200")
        forgot_username_frame.grid_columnconfigure(1, weight=1)

        label_forgot_username_last_name = customtkinter.CTkLabel(
            master=forgot_username_frame, text="Enter your last name here: ")
        label_forgot_username_last_name.grid(row=0, column=0, padx=10)
        entry_forgot_username_last_name = customtkinter.CTkEntry(
            master=forgot_username_frame)
        entry_forgot_username_last_name.grid(row=0, column=1, padx=10)

        label_forgot_username_email = customtkinter.CTkLabel(
            master=forgot_username_frame, text="Enter your email here : ")
        label_forgot_username_email.grid(row=1, column=0, padx=10)
        entry_forgot_username_email = customtkinter.CTkEntry(
            master=forgot_username_frame)
        entry_forgot_username_email.grid(row=1, column=1, padx=10)

        def find_my_username():
            print("finding username")
            lastname = entry_forgot_username_last_name.get()
            fu_email = entry_forgot_username_email.get()
            conn = sqlite3.connect('users.db')
            c = conn.cursor()

            c.execute(
                '''SELECT * FROM users WHERE last_name=? AND email=?''', (lastname, fu_email))
            result = c.fetchone()
            if result:
                messagebox.showinfo(
                    "Success", "Here is your username : " + str(result[4]))
                username_retreived = str(result[4])
                conn.close()
                forgot_username_frame.destroy()
                entry1.delete(0, 'end')
                entry1.insert(0, username_retreived)
                print("Username retreived")
                return
            else:
                messagebox.showerror(
                    "Error", "There is no Username related to this Last name and Email")
                print("No username for the informations entered, trying again")

        button_find_username = customtkinter.CTkButton(
            master=forgot_username_frame, text="Find my Username", command=find_my_username)
        button_find_username.grid(row=2, column=0, columnspan=2, pady=10)

    def forgot_password():
        print("forgot password")
        forgot_password_option_frame.destroy()
        forgot_password_frame = customtkinter.CTk()
        forgot_password_frame.title("Forgot Password")
        forgot_password_frame.geometry("400x200")
        forgot_password_frame.grid_columnconfigure(1, weight=1)

        label_forgot_password_last_name = customtkinter.CTkLabel(
            master=forgot_password_frame, text="Enter your username here: ")
        label_forgot_password_last_name.grid(row=0, column=0, padx=10)
        entry_forgot_password_last_name = customtkinter.CTkEntry(
            master=forgot_password_frame)
        entry_forgot_password_last_name.grid(row=0, column=1, padx=10)

        label_forgot_password_email = customtkinter.CTkLabel(
            master=forgot_password_frame, text="Enter your email here : ")
        label_forgot_password_email.grid(row=1, column=0, padx=10)
        entry_forgot_password_email = customtkinter.CTkEntry(
            master=forgot_password_frame)
        entry_forgot_password_email.grid(row=1, column=1, padx=10)

        def find_my_password():
            print("finding password")
            fp_username = entry_forgot_password_last_name.get()
            fp_email = entry_forgot_password_email.get()
            conn = sqlite3.connect('users.db')
            c = conn.cursor()

            c.execute(
                '''SELECT * FROM users WHERE username=? AND email=?''', (fp_username, fp_email))
            result = c.fetchone()
            if result:
                messagebox.showinfo(
                    "Success", "Here is your password : " + str(result[5]))
                password_retreived = str(result[5])
                conn.close()
                forgot_password_frame.destroy()
                entry2.delete(0, 'end')
                entry2.insert(0, password_retreived)
                print("Password retreived")
                return
            else:
                messagebox.showerror(
                    "Error", "There is no Password related to this Username and Email")
                print("No password for the informations entered, trying again")

        button_find_password = customtkinter.CTkButton(
            master=forgot_password_frame, text="Find my password", command=find_my_password)
        button_find_password.grid(row=2, column=0, columnspan=2, pady=10)

    label_forgot_username = customtkinter.CTkLabel(
        master=forgot_password_option_frame, text="Forgot Username?")
    label_forgot_username.grid(row=0, column=0, padx=10)
    forgot_username_button = customtkinter.CTkButton(
        master=forgot_password_option_frame, text="Clic here", command=forgot_username)
    forgot_username_button.grid(row=0, column=1, padx=10)

    forgot_password_label = customtkinter.CTkLabel(
        master=forgot_password_option_frame, text="Forgot Password?")
    forgot_password_label.grid(row=1, column=0, padx=10, pady=10)
    forgot_password_button = customtkinter.CTkButton(
        master=forgot_password_option_frame, text="Clic here", command=forgot_password)
    forgot_password_button.grid(row=1, column=1, pady=10, padx=10)


def Register():
    print("registration")
    Register_frame = customtkinter.CTk()
    Register_frame.title("Register")
    screen_width = Register_frame.winfo_screenmmwidth()
    screen_height = Register_frame.winfo_screenheight()

    # Calculating the x and y position to center the window
    x = int(screen_width/2 - 500/2)
    y = int(screen_height/2 - 350/2)
    Register_frame.geometry(f"400x200+{x}+{y}")
    Register_frame.grid_columnconfigure(1, weight=1)

    label_first_name = customtkinter.CTkLabel(
        master=Register_frame, text="First Name: ")
    label_first_name.grid(row=0, column=0, padx=10)
    first_name_entry = customtkinter.CTkEntry(master=Register_frame)
    first_name_entry.grid(row=0, column=1)

    label_last_name = customtkinter.CTkLabel(
        master=Register_frame, text="Last Name: ")
    label_last_name.grid(row=1, column=0, padx=10)
    last_name_entry = customtkinter.CTkEntry(master=Register_frame)
    last_name_entry.grid(row=1, column=1)

    label_email = customtkinter.CTkLabel(
        master=Register_frame, text="Email Address :")
    label_email.grid(row=2, column=0, padx=10)
    email_entry = customtkinter.CTkEntry(master=Register_frame)
    email_entry.grid(row=2, column=1)

    label_username = customtkinter.CTkLabel(
        master=Register_frame, text="Username: ")
    label_username.grid(row=3, column=0, padx=10)
    create_username_entry = customtkinter.CTkEntry(master=Register_frame)
    create_username_entry.grid(row=3, column=1)

    label_password = customtkinter.CTkLabel(
        master=Register_frame, text="Password: ")
    label_password.grid(row=4, column=0, padx=10)
    create_password_entry = customtkinter.CTkEntry(master=Register_frame)
    create_password_entry.grid(row=4, column=1)

    def store_credentials():
        print("Store credentials")
        first_name = first_name_entry.get()
        first_name = first_name_entry.get()
        last_name = last_name_entry.get()
        email = email_entry.get()
        username = create_username_entry.get()
        password = create_password_entry.get()

        conn = sqlite3.connect('users.db')
        c = conn.cursor()

        # Check if username already exists
        c.execute("SELECT * FROM users WHERE username=?", (username,))
        result = c.fetchone()
        if result:
            messagebox.showerror(
                "Error", "Username already exists. Please choose a different username.")
            print("Username already exist")
            return

        c.execute("SELECT * FROM users WHERE email=?", (email))
        if result:
            messagebox.showerror(
                "Error", "There is already an account for this email")
            return
        # ////////////////////////////////////////////////////////////////////
        # add something so someone with the same email cannot register 2 times
        # ////////////////////////////////////////////////////////////////////
        
        
        # Create table if it does not exist
        c.execute('''CREATE TABLE IF NOT EXISTS users
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                    first_name TEXT,
                    last_name TEXT,
                    email TEXT,
                    username TEXT,
                    password TEXT)''')

        # Insert user data into the table
        c.execute('''INSERT INTO users (first_name, last_name, email, username, password)
                    VALUES (?, ?, ?, ?, ?)''', (first_name, last_name, email, username, password))

        conn.commit()

        Register_frame.destroy()
        messagebox.showinfo("Success", "Registration Successful!")
        print("Registration made")

    customtkinter.set_default_color_theme("green")
    register_button = customtkinter.CTkButton(
        master=Register_frame, text="Register", command=store_credentials)
    register_button.grid(row=5, column=0, pady=10, padx=10)
    customtkinter.set_default_color_theme("dark-blue")


def open_new_interface():
    print("New_Interface_Stock_Search:")
    frame2 = customtkinter.CTk()
    frame2.title("Stock Research")
    screen_width = frame2.winfo_screenwidth()
    screen_height = frame2.winfo_screenheight()
    frame2.geometry(
        f"{int(screen_width * 0.9)}x{int(screen_height * 0.9)}+{int(screen_width * 0.05)}+{int(screen_height * 0.05)}")
    label = customtkinter.CTkLabel(
        master=frame2, text="Stocks Data Finder", font=("Arial", 30))
    label.pack(pady=20, padx=18)

    label_ticker = customtkinter.CTkLabel(
        master=frame2, text="Enter the Ticker of your stock under here:", font=("Arial", 18))
    label_ticker.pack(pady=0, padx=0)
    ticker_entry = customtkinter.CTkEntry(
        master=frame2, placeholder_text="Ticker", width=300)
    ticker_entry.pack(pady=0, padx=0)

    data_checkbox = customtkinter.BooleanVar()
    price_target_checkbox = customtkinter.BooleanVar()
    recommendations_checkbox = customtkinter.BooleanVar()

    checkbox1 = customtkinter.CTkCheckBox(
        master=frame2, text="________Data__________", variable=data_checkbox)
    checkbox1.pack(pady=15, padx=25)

    checkbox2 = customtkinter.CTkCheckBox(
        master=frame2, text="__Analysts price target___", variable=price_target_checkbox)
    checkbox2.pack(pady=15, padx=25)

    checkbox3 = customtkinter.CTkCheckBox(
        master=frame2, text="___Recommendations___", variable=recommendations_checkbox)
    checkbox3.pack(pady=15, padx=25)

    label_period = customtkinter.CTkLabel(
        master=frame2, text="Choose a period from these choices: \n 1mo, 3mo, 6mo, ytd, 1y, 2y, 5y, 10y, max", font=("Arial", 18))
    label_period.pack(pady=5, padx=5)
    entry3 = customtkinter.CTkEntry(master=frame2, placeholder_text="Period")
    entry3.pack(pady=5, padx=10)

    def results_interface():
        print("Results interface for : " + ticker_entry.get())
        frame3 = tk.Toplevel()
        frame3.title("Results for: "+ticker_entry.get())
        screen_width = frame3.winfo_screenwidth()
        screen_height = frame3.winfo_screenheight()
        frame3.geometry(
            f"{int(screen_width * 0.9)}x{int(screen_height * 0.9)}+{int(screen_width * 0.05)}+{int(screen_height * 0.05)}")
        label = customtkinter.CTkLabel(
            master=frame3, text="Results for "+ticker_entry.get(), font=("Arial", 24))
        label.pack(pady=12, padx=10)

        # create a treeview with 7 columns
        tree = ttk.Treeview(frame3, columns=(
            'Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Dividends', 'Stock Splits'))
        style = ttk.Style()
        style.theme_use("clam")

        # set column headings
        tree.heading('#0', text='Ticker', anchor='center')
        tree.heading('#1', text='Date', anchor='center')
        tree.heading('#2', text='Open', anchor='center')
        tree.heading('#3', text='High', anchor='center')
        tree.heading('#4', text='Low', anchor='center')
        tree.heading('#5', text='Close', anchor='center')
        tree.heading('#6', text='Volume', anchor='center')
        tree.heading('#7', text='Dividends', anchor='center')
        tree.heading('#8', text='Stock Splits', anchor='center')

        # set column widths
        tree.column('#1', width=100)
        tree.column('#2', width=100)
        tree.column('#3', width=100)
        tree.column('#4', width=100)
        tree.column('#5', width=100)
        tree.column('#6', width=100)
        tree.column('#7', width=100)
        tree.column('#8', width=100)

        # insert data into the treeview
        if data_checkbox.get():
            ticker_symbol = ticker_entry.get()
            ticker = yf.Ticker(ticker_symbol)
            # select the period
            history = ticker.history(period=entry3.get())
            for i, row in history.iterrows():
                open_price = round(row['Open'], 2)
                high_price = round(row['High'], 2)
                low_price = round(row['Low'], 2)
                close_price = round(row['Close'], 2)
                volume = round(row['Volume'], 2)
                dividends = round(row['Dividends'], 2)
                stock_splits = round(row['Stock Splits'], 2)

                # insert the data into the treeview
                tree.insert('', 'end', text=ticker.ticker, values=(i.date(
                ), f"{open_price:^10}", f"{high_price:^10}", f"{low_price:^10}", f"{close_price:^10}", f"{volume:^10}", f"{dividends:^10}", f"{stock_splits:^10}"))

            # pack the treeview into the root window
            tree.pack(fill='both', expand=True, padx=10, side='left')

            # create a vertical scrollbar for the treeview
            vsb = ttk.Scrollbar(frame3, orient="vertical", command=tree.yview)
            vsb.pack(side='left', fill='y')

            # configure the treeview to use the scrollbar
            tree.configure(yscrollcommand=vsb.set)

            # set the height of the treeviw to show more rows
            tree['height'] = 20

        def export_to_excel(tree):
            wb = openpyxl.Workbook()
            ws = wb.active

            # write column headings
            for col, heading in enumerate(tree["columns"]):
                ws.cell(row=1, column=col+1,
                        value=tree.heading(heading)["text"])

            # write data
            for row, item in enumerate(tree.get_children()):
                for col, value in enumerate(tree.item(item)["values"]):
                    ws.cell(row=row+2, column=col+1, value=value)

            # save the workbook
            wb.save(ticker_symbol+"_treeview_data.xlsx")
            print("Exported to Excel")

        export_button = tk.Button(
            frame3, text="Export to Excel", command=lambda: export_to_excel(tree))
        export_button.pack(side='right', padx=10, pady=10)

        frame3.mainloop()

    button2 = customtkinter.CTkButton(
        master=frame2, text="Search", command=results_interface)
    button2.pack(pady=30, padx=10)

    def exit_program():
        frame2.destroy()
    customtkinter.set_default_color_theme("green")
    exit_button = customtkinter.CTkButton(
        master=frame2, text="Close the program", command=exit_program)
    exit_button.pack(pady=100)
    customtkinter.set_default_color_theme("dark-blue")
    print("Program closed")

    frame2.mainloop()


frame = customtkinter.CTkFrame(master=root)
frame.pack(pady=20, padx=60, fill="both", expand=True)

label = customtkinter.CTkLabel(
    master=frame, text="Login System", font=("Arial", 24))
label.pack(pady=12, padx=10)

entry1 = customtkinter.CTkEntry(master=frame, placeholder_text="Username")
entry1.pack(pady=12, padx=10)

entry2 = customtkinter.CTkEntry(
    master=frame, placeholder_text="Password", show="*")
entry2.pack(pady=12, padx=10)

button = customtkinter.CTkButton(master=frame, text="Login", command=login)
button.pack(pady=12, padx=10)

customtkinter.set_default_color_theme("green")
Not_register_button = customtkinter.CTkButton(
    master=frame, text="Sign-up Now", command=Register)
Not_register_button.pack(pady=12, padx=10)
forgot_button = customtkinter.CTkButton(
    master=frame, text="Forgot credentials?", command=forgot_credentials)
forgot_button.pack(padx=10, pady=10)
customtkinter.set_default_color_theme("dark-blue")

checkbox_var = customtkinter.BooleanVar()
checkbox = customtkinter.CTkCheckBox(
    master=frame, text="Remember me", variable=checkbox_var)
checkbox.pack(pady=12, padx=10)

optionmenu_var = customtkinter.StringVar(value="System")


def optionmenu_callback(choice):
    print("optionmenu dropdown clicked:", choice)
    if choice == "System":
        customtkinter.set_appearance_mode("system")
    if choice == "Light":
        customtkinter.set_appearance_mode("light")
    elif choice == "Dark":
        customtkinter.set_appearance_mode("dark")


combobox = customtkinter.CTkOptionMenu(master=frame,
                                       values=["System", "Light", "Dark"],
                                       command=optionmenu_callback,
                                       variable=optionmenu_var)
combobox.pack(padx=20, pady=10)

load_credentials()

root.mainloop()
